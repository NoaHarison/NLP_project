import zipfile
import os
import pandas as pd
from openpyxl import Workbook
import re
from nltk.corpus import words
from collections import Counter
from decimal import Decimal, ROUND_HALF_UP

from google.colab import drive
drive.mount('/content/drive', force_remount=True)

english_words = set(words.words())

def is_english_word(word):
    word = re.sub(r'[^a-zA-Z]', '', word)
    return all(part.lower() in english_words for part in word.split())

def calculate_english_ttr(text):
    # Remove punctuation marks
    text = re.sub(r'[^\w\s]', '', text)
    # Split text into words, keeping hyphenated words together
    words = re.findall(r'\w+(?:-\w+)*', text)
    # Filter out non-English words
    english_words = [word for word in words if is_english_word(word)]
    # Count unique English words
    unique_english_words = len(set(english_words))
    # Count total English words
    total_english_words = len(english_words)
    # Calculate TTR for English words
    if total_english_words == 0:
        english_ttr = Decimal('0')
    else:
        english_ttr = Decimal(unique_english_words) / Decimal(total_english_words)
        english_ttr = english_ttr.quantize(Decimal('.01'), rounding=ROUND_HALF_UP)
    return f"Type-Token Ratio for English words: {english_ttr}"

# Path to the ZIP file
zip_path = 'drive/MyDrive/bioinformatica/Colab_Notebooks/final_project/basic_analayze_data/record_kids_data/data.zip'
# List to store file contents and number sequences
file_contents = []

# Open the ZIP file
with zipfile.ZipFile(zip_path, 'r') as zip_ref:
    # Loop through each file in the ZIP
    for file_info in zip_ref.infolist():
        # Check if the file is a text file
        if 'finished' in file_info.filename and 'Brown/Sarah' in file_info.filename:
            # Read the file contents
            with zip_ref.open(file_info.filename) as file:
                contents = file.read().decode('utf-8')
            # Split the contents into sentences (each line is a sentence)
            sentences = contents.split('\\n')
            # Get the number sequence from the file name
            file_nums = ''.join(filter(str.isdigit, file_info.filename))
            # Calculate TTR for each sentence
            ttr_values = [calculate_english_ttr(sentence) for sentence in sentences]
            # Add sentences and TTR values to the list
            file_contents.extend([(sentence, file_nums, ttr) for sentence, ttr in zip(sentences, ttr_values)])

# Create a new Excel file
workbook = Workbook()
worksheet = workbook.active
# Write column headers
worksheet['A1'] = 'Text'
worksheet['B1'] = 'Number Sequence'
worksheet['C1'] = 'TTR'
# Write data to the worksheet
for row, (text, file_nums, ttr) in enumerate(file_contents, start=2):
    worksheet.cell(row=row, column=1, value=text)
    worksheet.cell(row=row, column=2, value=file_nums)
    worksheet.cell(row=row, column=3, value=ttr)

# Save the Excel file
excel_file_path = 'drive/MyDrive/bioinformatica/Colab_Notebooks/final_project/basic_analayze_data.xlsx'
workbook.save(excel_file_path)
print(f"Excel file saved successfully: {excel_file_path}")

import pandas as pd
import matplotlib.pyplot as plt

# read data
data_df = pd.read_excel('drive/MyDrive/bioinformatica/Colab_Notebooks/final_project/basic_analayze_data/length_and_non_english/Sarah_data_analysis_result_formatted.xlsx')

# create scatter
plt.figure(figsize=(10, 6))
plt.scatter(data_df['age'], data_df['TTR'], color='blue', alpha=0.5)
plt.title('Sarah: TTR vs. Age')
plt.xlabel('Age')
plt.ylabel('TTR')
plt.grid(True)
plt.show()




english_words = set(words.words())

def is_english_word(word):
    word = re.sub(r'[^a-zA-Z]', '', word)
    return word.lower() in english_words

# Path to the ZIP file
zip_path = 'drive/MyDrive/bioinformatica/Colab_Notebooks/final_project/basic_analayze_data/record_kids_data/data.zip'
# List to store file contents and number of unique English words
file_contents = []

# Open the ZIP file
with zipfile.ZipFile(zip_path, 'r') as zip_ref:
    # Loop through each file in the ZIP
    for file_info in zip_ref.infolist():
        # Check if the file is a text file
        if 'finished' in file_info.filename and 'Brown/Sarah' in file_info.filename:
            # Read the file contents
            with zip_ref.open(file_info.filename) as file:
                contents = file.read().decode('utf-8')
            # Split the contents into sentences (each line is a sentence)
            sentences = contents.split('\\n')
            # Get the number sequence from the file name
            file_nums = ''.join(filter(str.isdigit, file_info.filename))
            # Filter out non-English words from each sentence
            english_words_list = []
            for sentence in sentences:
                words = re.findall(r'\w+(?:-\w+)*', sentence)
                english_words_list.extend([word for word in words if is_english_word(word)])
            # Count unique English words
            unique_english_words = len(set(english_words_list))
            # Add the number of unique English words to the list
            file_contents.append((file_nums, unique_english_words))

# Create a new Excel file
workbook = Workbook()
worksheet = workbook.active
# Write column headers
worksheet['A1'] = 'Number Sequence'
worksheet['B1'] = 'Unique English Words'
# Write data to the worksheet
for row, (file_nums, unique_english_words) in enumerate(file_contents, start=2):
    worksheet.cell(row=row, column=1, value=file_nums)
    worksheet.cell(row=row, column=2, value=unique_english_words)

# Save the Excel file
excel_file_path = 'drive/MyDrive/bioinformatica/Colab_Notebooks/final_project/basic_analayze_data.xlsx'
workbook.save(excel_file_path)
print(f"Excel file saved successfully: {excel_file_path}")
